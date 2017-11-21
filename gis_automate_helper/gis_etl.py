import openpyxl, os, arcpy, codecs, csv


def read_excel(excel_file, excel_worksheet_name):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.get_sheet_by_name(excel_worksheet_name)

    xls_keys = []
    i = 1
    while i < ws.max_column + 1:
        xls_keys.append(ws.cell(row=1, column=i).value)
        i += 1

    xls_dict = []
    j = 2
    while j < ws.max_row + 1:
        i = 1
        d = None
        while i < ws.max_column + 1:
            cell_value_class = ws.cell(row=1, column=i).value
            cell_value_id = ws.cell(row=j, column=i).value
            if d is None:
                d = {cell_value_class: cell_value_id}
            else:
                d[cell_value_class] = cell_value_id
            i += 1
        xls_dict.append(d)
        j += 1

    return xls_dict


def read_csv(csv_file):
    open_csv = file(csv_file)
    arry = csv.DictReader(open_csv)

    return arry


def create_excel(outfolder, workbookname):
    if ".xlsx" not in workbookname:
        workbookname += ".xlsx"

    fname = os.path.join(outfolder, workbookname)

    if os.path.exists(fname):
        os.remove(fname)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(outfolder, workbookname))

    return wb


def create_csv(outfolder, csvname):
    if ".csv" not in csvname:
        csvname += ".csv"

    fname = os.path.join(outfolder, csvname)

    if os.path.exists(fname):
        os.remove(fname)

    csvfile = codecs.open(fname, 'w', 'utf-8')

    return csvfile


def export_feature_class_to_excel(outfolder, workbookname, header_fields, fc, sql):
    wb = create_excel(outfolder=outfolder, workbookname=workbookname)
    ws = wb.active

    colcnt = 1
    for fld in header_fields:
        ws.cell(row=1, column=colcnt).value = fld
        colcnt += 1

    rowcount = 2
    for row in arcpy.da.SearchCursor(fc, header_fields, sql):
        colcnt = 1
        for fld in row:
            if fld is not None:
                ws.cell(row=rowcount, column=colcnt).value = fld
            colcnt += 1
        rowcount += 1

    wb.save(os.path.join(outfolder, workbookname))

    del ws
    del wb


def export_array_to_excel(outfolder, workbookname, header_fields, out_array):
    wb = create_excel(outfolder=outfolder, workbookname=workbookname)
    ws = wb.active

    colcnt = 1
    for fld in header_fields:
        ws.cell(row=1, column=colcnt).value = fld
        colcnt += 1

    rowcount = 2
    for row in out_array:
        colcnt = 1
        for fld in row:
            if fld is not None:
                ws.cell(row=rowcount, column=colcnt).value = fld
            colcnt += 1
        rowcount += 1

    wb.save(os.path.join(outfolder, workbookname))

    del ws
    del wb


def export_array_to_csv(outfolder, csvname, header, out_array):
    csvfile = create_csv(outfolder=outfolder, csvname=csvname)

    outstring = ""
    for field in header:
        outstring += "\"" + field + "\","

    csvfile.write(outstring[:-1])
    csvfile.write("\"\n")

    for row in out_array:
        outstring = ""
        for field in row:
            outstring += "\"" + field + "\","
        csvfile.write(outstring[:-1])
        csvfile.write("\"\n")

    csvfile.close()

    del csvfile
